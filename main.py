import openpyxl
from openpyxl import load_workbook, Workbook

# Путь к Excel-файлу
FILE_PATH = "students_data.xlsx"

# Инициализация файла Excel при первом запуске
def initialize_excel():
    try:
        workbook = load_workbook(FILE_PATH)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append(["ФИО", "Группа", "Курс", "Средний балл"])
        workbook.create_sheet(title="Averages")
        averages = workbook["Averages"]
        averages.append(["Группа", "Курс", "Средний балл по группе/курсу"])
        workbook.save(FILE_PATH)

# Функция для отображения всех студентов
def display_students():
    workbook = load_workbook(FILE_PATH)
    sheet = workbook["Data"]

    print("\nСписок студентов:")
    print("№\tФИО\t\tГруппа\tКурс\tСредний балл")
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        print(f"{i}\t{row[0]}\t{row[1]}\t{row[2]}\t{row[3]:.2f}")
    workbook.close()

# Функция для выбора студента по ФИО или номеру
def select_student():
    display_students()
    choice = input("\nВыберите метод поиска студента:\n1 - По номеру\n2 - По ФИО\nВведите выбор: ")
    
    workbook = load_workbook(FILE_PATH)
    sheet = workbook["Data"]
    selected_student = None
    student_index = None
    
    if choice == "1":  # Поиск по номеру
        try:
            number = int(input("Введите номер студента: "))
            if number < 1 or number > sheet.max_row - 1:
                print("Ошибка: Неверный номер студента.")
            else:
                student_index = number + 1
                selected_student = list(sheet.iter_rows(min_row=student_index, max_row=student_index, values_only=True))[0]
        except ValueError:
            print("Ошибка: Введите корректный номер студента.")
    
    elif choice == "2":  # Поиск по ФИО
        name = input("Введите ФИО студента: ")
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == name:
                selected_student = row
                student_index = i
                break
        if not selected_student:
            print("Ошибка: Студент с таким ФИО не найден.")
    
    else:
        print("Ошибка: Некорректный выбор.")

    workbook.close()
    return selected_student, student_index

# Функция для добавления оценок студенту
def add_grades():
    student, index = select_student()
    if not student:
        return

    workbook = load_workbook(FILE_PATH)
    sheet = workbook["Data"]
    print(f"\nДобавление оценок для студента: {student[0]}")
    try:
        grades = list(map(int, input("Введите оценки через пробел (по 100-балльной шкале): ").split()))
        avg_grade = sum(grades) / len(grades)
        
        # Вставляем оценки в столбцы начиная с 5-го
        for i, grade in enumerate(grades, start=5):
            sheet.cell(row=index, column=i, value=grade)
        
        # Обновляем средний балл
        sheet.cell(row=index, column=4, value=avg_grade)
        print("Оценки успешно добавлены!")
    except ValueError:
        print("Ошибка: Введите числовые значения.")
    
    workbook.save(FILE_PATH)
    workbook.close()

# Функция для подсчета средних баллов по группам и курсам
def calculate_averages():
    workbook = load_workbook(FILE_PATH)
    sheet = workbook["Data"]
    averages = workbook["Averages"]
    averages.delete_rows(2, averages.max_row)  # Очищаем старые записи

    group_course = {}
    total_sum, total_count = 0, 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        group, course, avg = row[1], row[2], row[3]
        key = (group, course)
        if key not in group_course:
            group_course[key] = []
        group_course[key].append(avg)

        total_sum += avg
        total_count += 1

    for (group, course), avgs in group_course.items():
        overall_avg = sum(avgs) / len(avgs)
        averages.append([group, course, overall_avg])

    # Добавляем общий средний балл
    if total_count > 0:
        overall_total_avg = total_sum / total_count
        averages.append(["-", "-", "Средний балл среди всех студентов", overall_total_avg])
    workbook.save(FILE_PATH)
    workbook.close()
    print("Средние баллы успешно рассчитаны и сохранены!")

# Основное меню
def main():
    initialize_excel()
    while True:
        print("\n=== Система обработки аттестации студентов ===")
        print("1. Внести оценки студенту")
        print("2. Посмотреть данные студента")
        print("3. Рассчитать средние баллы")
        print("4. Выход")
        choice = input("Выберите действие (1-4): ")
        
        if choice == "1":
            add_grades()
        elif choice == "2":
            student, _ = select_student()
            if student:
                print("\nДанные студента:")
                print(f"ФИО: {student[0]}\nГруппа: {student[1]}\nКурс: {student[2]}\nСредний балл: {student[3]:.2f}")
        elif choice == "3":
            calculate_averages()
        elif choice == "4":
            print("Выход из программы...")
            break
        else:
            print("Ошибка: Некорректный ввод. Попробуйте снова.")

if __name__ == "__main__":
    main()
